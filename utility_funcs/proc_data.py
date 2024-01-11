import pickle
import os

import numpy as np
import pandas as pd
import networkx as nx
from scipy import interpolate
from openpyxl import load_workbook


def safe_dump_obj(obj, fname):
    """Dump the object to pickle file fname. Create parent directory if needed.

    Args:
        obj (Anything): the python object to dump.
        fname (str): the pickle file's full path name.
    """
    if fname is None or obj is None:
        return
    os.makedirs(os.path.dirname(fname), exist_ok=True)
    with open(fname, "wb") as f:
        pickle.dump(obj, f)


def aggregate(a, n=3):
    cumsum = np.cumsum(a, dtype=float)
    ret = []
    for i in range(-1, len(a) - n, n):
        low_index = i if i >= 0 else 0
        ret.append(cumsum[low_index + n] - cumsum[low_index])
    return ret


def load_data(
    file_path,
    sheet_name="Sheet1",
    aggre_delta=1,
    normalize=True,
    zero_fill_method="prevlatter",
    verbose=True,
):
    """Load metric data from file_path. Each column is one variable.

    Params:
        file_path:
        sheet_name: name of sheet to load.
        normaliza: normalize data by subtract mean and divide by std.
        fill_zeros: fill 0 data with nearest sample.
        verbose: the debugging print level: 0 (Nothing), 1 (Method info), 2 (Phase info), 3(Algorithm info)

    Returns:
        data     : data in numpy array format, shape of [T, N], each column is a variable
        data_head: service names
    """
    # verbose >= 3, print data loading info
    if verbose and verbose >= 3:
        print("{:^10}{:<30}:".format("", "Data path"), file_path)
    # region Read excel sheet, each row of data is one variable
    wb = load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    data = []
    data_head = []
    for row_values in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True
    ):
        data_head.append(row_values[0])
        data.append(row_values[1:])

    if verbose and verbose >= 3:
        print("{:^10}{:<30}: ".format("", "Sheet Names"), end="")
        for name in wb.sheetnames:
            print(name, end=", ")
        print("")
        print("{:^10}{:<30}:".format("", "Loaded Sheet"), sheet_name)
        print(
            "{:^10}{:<30}:".format("", "Sheet Size"),
            "{} x {}".format(sheet.max_row, sheet.max_column),
        )
    wb.close()
    # endregion

    # region Aggregate data
    if aggre_delta != 1:
        # Aggregate data
        data = [aggregate(row, aggre_delta) for row in data]
    # transpose data, now each column is one variable
    data = np.array(data).T
    if verbose and verbose >= 3:
        print("{:^10}{:<30}:".format("", "Aggregate delta"), aggre_delta)
        print("{:^10}{:<30}:".format("", "Data Shape"), data.shape)
    # endregion

    zero_count = np.sum(data == 0, axis=0)
    # Fill 0s in data
    if zero_fill_method == "prevlatter":
        if verbose:
            print(
                "{:^10}{:<30}:".format("", "Zero fill method"), "Previous then latter"
            )
        filled_data = data.copy()
        for j in range(filled_data.shape[1]):
            for i in range(filled_data.shape[0]):
                if filled_data[i, j] == 0 and i >= 1:
                    filled_data[i, j] = filled_data[i - 1, j]
        for j in range(filled_data.shape[1] - 1, -1, -1):
            for i in range(filled_data.shape[0] - 1, -1, -1):
                if filled_data[i, j] == 0 and i <= filled_data.shape[0] - 2:
                    filled_data[i, j] = filled_data[i + 1, j]
        data = filled_data
    elif zero_fill_method in [
        "linear",
        "nearest",
        "zero",
        "slinear",
        "quadratic",
        "cubic",
        "previous",
        "next",
    ]:
        # Possible interpolate methods are:
        # linear, nearest, zero, slinear, quadratic, cubic, previous, next
        if verbose:
            print(
                "{:^10}{:<30}:".format("", "Zero fill method"),
                zero_fill_method + " interpolate",
            )
        x = np.arange(data.shape[0])
        new_data = []
        for var in range(data.shape[1]):
            ind = data[:, var].nonzero()
            f = interpolate.interp1d(
                x[ind[0]],
                data[ind[0], var],
                kind=zero_fill_method,
                fill_value="extrapolate",
            )
            new_data.append(f(x))
        data = np.array(new_data).T

    # Normalize data by subtract mean and divide by std
    if normalize:
        data_mean = np.mean(data, axis=0, keepdims=True)
        data_std = np.std(data, axis=0, keepdims=True)
        data = (data - data_mean) / data_std

    # print data attributes
    if verbose:
        print("{:^10}{:<30}:".format("", "Data header"))
        for i, head in enumerate(data_head):
            print("{:>15}({:4d} 0s):{}".format(i + 1, zero_count[i], head))
    return data, data_head


tcdf_data_path = "data/"
timeseries_files = [
    "manyinputs_returns30007000_header.csv",
    "random-rels_20_1A_returns30007000_header.csv",
    "random-rels_20_1B_returns30007000_header.csv",
    "random-rels_20_1C_returns30007000_header.csv",
    "random-rels_20_1D_returns30007000_header.csv",
    "random-rels_20_1E_returns30007000_header.csv",
    "random-rels_40_1_returns30007000_header.csv",
    "random-rels_40_1_3_returns30007000_header.csv",
    "random-rels_20_1_3_returns30007000_header.csv",
]
gt_files = [
    "manyinputs.csv",
    "random-rels_20_1A.csv",
    "random-rels_20_1B.csv",
    "random-rels_20_1C.csv",
    "random-rels_20_1D.csv",
    "random-rels_20_1E.csv",
    "random-rels_20_1_3.csv",
    "random-rels_40_1.csv",
    "random-rels_40_1_3.csv",
]

data_store = {"finance": (timeseries_files, gt_files)}


def load_tcdf_data(dir="finance"):
    """Load the TCDF paper's sythetic time series data with ground truth causal graph.

    Params:
        dir: which datasets to load. Can be "finance" or "fmri". See more in the paper.
    Returns:
        datasets: A list of DataFrames containing the datasets. Each DataFrame's shape is [T, N],
            where T is the number of data samples and N is the number of variables.
        gt_graphs: A list of ground truth causal graphs. Edge also has a 'lag' attribute indicating
            the causal delays.
    """
    datasets = []
    gt_graphs = []
    if dir in data_store:
        for ts_f, gt_f in zip(*data_store[dir]):
            data = pd.read_csv(os.path.join(tcdf_data_path, "Finance", ts_f), header=0)
            N = len(data.columns)
            gt_graph = nx.DiGraph()
            gt_graph.add_nodes_from(range(N))
            with open(os.path.join(tcdf_data_path, "Finance", gt_f), "rt") as f:
                for line in f:
                    s, e, lag = line.strip().split(",")
                    gt_graph.add_edge(int(s), int(e), lag=lag)
            datasets.append(data)
            gt_graphs.append(gt_graph)
    elif dir == "fmri":
        for i in range(1, 29, 1):
            data = pd.read_csv(
                os.path.join(tcdf_data_path, "fMRI", f"timeseries{i}.csv"), header=0
            )
            N = len(data.columns)
            gt_graph = nx.DiGraph()
            gt_graph.add_nodes_from(range(N))
            with open(
                os.path.join(tcdf_data_path, "fMRI", f"sim{i}_gt_processed.csv"), "rt"
            ) as f:
                for line in f:
                    s, e, lag = line.strip().split(",")
                    gt_graph.add_edge(int(s), int(e), lag=lag)
            datasets.append(data)
            gt_graphs.append(gt_graph)
    else:
        print("No such datasets.")
        exit(1)
    return datasets, gt_graphs


def load_pairs_data():
    """Load the pairs dataset from https://webdav.tuebingen.mpg.de/cause-effect/.
    Returns:
        dfs: A list of DataFrames containing the datasets. Each DataFrame's shape is [T, N],
            where T is the number of data samples and N is the number of variables.
        relations: A list of tuples (cause, effect). cause or effect could be 'x' or 'y'.
            'x' is the first column and 'y' is the second column.
    """
    data_root = 'data/pairs'

    idx = 1
    dfs = []
    relations = []
    for idx in range(1, 109):
        # Problem data.
        if idx in [52, 54, 55, 71, 72, 81, 82, 83, 86, 105]:
            continue
        data_fname = os.path.join(data_root, "pair{:04}.txt".format(idx))
        desc_fname = os.path.join(data_root, "pair{:04}_des.txt".format(idx))
        # print(data_fname)
        a = np.genfromtxt(data_fname)
        df = pd.DataFrame(a)
        with open(desc_fname, "rt") as f:
            lines = f.readlines()
        lines = [line.lower() for line in lines]
    
        def findline(s):
            for line in lines:
                for i in s:
                    if line.find(i) != -1:
                        ret = line.split(i)[1].strip()
                        return ret
            return None
        x_line = findline(["x:", "x =", "(x):"])
        y_line = findline(["y:", "y =", "(y):"])
        colnames = []
        
        if x_line is not None and '\t' in x_line:
            colnames.extend(x_line.split('\t'))
        elif x_line is not None:
            colnames.append(x_line)
        if y_line is not None and '\t' in y_line:
            colnames.extend(y_line.split('\t'))
        elif y_line is not None:
            colnames.append(y_line)
        if df.shape[1] == len(colnames):
            df.columns = colnames
            # print(df.columns)
        elif df.shape[1] > len(colnames):
            print("Not enough column names in data description!")
            break
        else:
            print("Too many column names in data description!")
            break
        
        gt_str = None
        if 'ground truth:\n' in lines:
            gt_str = lines[lines.index('ground truth:\n')+1].strip()
            # print(gt_str)
        else:
            for line in lines:
                if '-->' in line:
                    gt_str = line.strip()
                    # print(gt_str)
                    break
                elif 'ground truth' in line:
                    gt_str = line.lstrip('ground truth').strip(': \n')
                    # print(gt_str)
                    break
        if gt_str is None:
            print("No groundtruth found!")
            break
        if '-->' in gt_str:
            l, r = gt_str.split('-->')
        elif '->' in gt_str:
            l, r = gt_str.split('->')
        l = l.strip()
        r = r.strip()
        
        relation = (l, r)
        dfs.append(df)
        relations.append(relation)
    return dfs, relations